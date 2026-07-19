"""Dump every sheet of the workbook: formulas, cached values, validations, names."""
import json
import sys

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from paths import WORKBOOK, WORKBOOK_DUMP
src, dst = str(WORKBOOK), str(WORKBOOK_DUMP)
wbf = openpyxl.load_workbook(src, data_only=False)
wbv = openpyxl.load_workbook(src, data_only=True)

info = {"defined_names": {}, "sheets": []}
try:
    for name, dn in wbf.defined_names.items():
        info["defined_names"][name] = dn.value
except AttributeError:
    for dn in wbf.defined_names.definedName:
        info["defined_names"][dn.name] = dn.value

for wsf in wbf.worksheets:
    wsv = wbv[wsf.title]
    sheet = {
        "title": wsf.title,
        "state": wsf.sheet_state,
        "dims": [wsf.max_row, wsf.max_column],
        "merged": [str(r) for r in wsf.merged_cells.ranges],
        "validations": [
            {"sqref": str(dv.sqref), "type": dv.type, "formula1": dv.formula1,
             "formula2": dv.formula2}
            for dv in wsf.data_validations.dataValidation
        ],
        "cells": [],
    }
    for row in wsf.iter_rows():
        for c in row:
            v = c.value
            cached = wsv[c.coordinate].value
            if v is None and cached is None:
                continue
            entry = {"a": c.coordinate}
            if isinstance(v, ArrayFormula):
                entry["f"] = v.text
                entry["array"] = True
            elif isinstance(v, str) and v.startswith("="):
                entry["f"] = v
            elif v is not None:
                entry["v"] = str(v) if hasattr(v, "isoformat") else v
            if "f" in entry or entry.get("v") != cached:
                if cached is not None:
                    entry["cached"] = str(cached) if hasattr(cached, "isoformat") else cached
            if c.number_format and c.number_format != "General":
                entry["nf"] = c.number_format
            sheet["cells"].append(entry)
    info["sheets"].append(sheet)

with open(dst, "w", encoding="utf-8") as fh:
    json.dump(info, fh, indent=1, default=str, ensure_ascii=False)

for s in info["sheets"]:
    print(s["title"], s["state"], s["dims"], len(s["cells"]), "cells,",
          len(s["validations"]), "validations")
print("defined names:", len(info["defined_names"]))
